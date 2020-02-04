#Проверим как работает вставка строки на объёденённые ячейки
ws.merge_cells('D4:E8')
# Добавить новую строку с данными на лист ws в конец листа
ws.append([1,2,3,41,51,6,7,8,9,10,11,12,13,14,15])# как видно часть данных пропала и вставка строки пошла на следующкю, так как запись происходит в конец файла, ,ошибки не было
# Добавить новую строку с данными на лист ws в конец листа
ws.append([1,2,3,42,52,6,7,8,9,10,11,12,13,14,15])# тут тоже пропала часть данных,ошибки не было
ws.insert_rows(6) #вставить строку на 6 позицию, все данные сохронятся вставится пустая строка
ws.append([1,22,3,42,52,6,7,8,9,10,11,12,13,14,15]) # но данные вставятся только через 2 строки, в общем кривоватая библиотека немного, но это не беда
#удалит столбец с данными
ws.delete_cols(5) #объяденённые ячейки не тронит
#Вывод: лучще повторно открывать документ для корректировки строк и столбцов, а для вставки данных также

#Сгрупировать столбцы
ws.column_dimensions.group('A','D')
#Сгрупировать и скрыть строки
ws.row_dimensions.group(1,10, hidden=True)
ws['A']
#Создать новый лист в книге
ws2 = wb.create_sheet("test_worksheet_2", 1)# пораметр, показывает где в списке листов создать лист
#Присвоить цвет значку листа
ws2.sheet_properties.tabColor = "1072BA"#hex код цвета
#Выводим список имён листов книги
print(wb.sheetnames)
#Копировать лист
ws3=wb.copy_worksheet(ws)
#Меняем название листа
ws3.title="test_copy"
#Копировать лист 2
ws4=wb.copy_worksheet(ws2)# По кмолчанию добавляется " Copy" к названию листа
#Копировать лист 2 повторно и удалим
ws5=wb.copy_worksheet(ws2)
print(wb.sheetnames)
#Удалить лист
wb.remove(ws5)
print(wb.sheetnames)
# Поменять листы местами
# выбираем имена
a=wb.sheetnames
#определяем порядок следования листов
wb._sheets =[wb[a[1]],wb[a[0]],wb[a[2]],wb[a[3]]]
#Сохранить книгу по имени
wb.save("test_workbook.xlsx")
#Закрыть книгу
wb.close()
'''Работа со стилями, фильтрами'''
#Открытие существкющей книги    
#Открыть книгу по имени
wb = openpyxl.load_workbook('test_workbook.xlsx')#Для быстроты чтения при работе с данными лучше ставить аргумент read_only=True, но тогда многие переменные не используются
ws=wb["test_copy"]

# Выбрать область под фильтрацию и сортировку
ws.auto_filter.ref = "A1:N10"
ws.auto_filter.add_filter_column(0, [1]) #Добавляем фильтр про списку значений, пробывал доавляется только значёк
ws.auto_filter.add_sort_condition("C2:C10") #Делаем значёк сортировки, пробывал добавляется только значёк
#проблемма реuается перестоновкой данных столбцов и скрытием строк, например так
ws.column_dimensions['B'].hidden= True
wb.save("test_workbook1.xlsx")
wb.close()



from openpyxl.utils import FORMULAE# множество формул
#переведём в кортеж
a=tuple(FORMULAE)
#вывидим часть получившегося кортежа
for i in range(10):
    print(a[i])
